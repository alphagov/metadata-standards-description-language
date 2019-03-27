import openpyxl as opxl
from operator import itemgetter

toclst = []
colhead = []
datalst = []

# This will get the file, initial we have a default file, later we will allow them to input
def getfile():
    wb = opxl.load_workbook('./Excel Import Text.xlsx')
    # wb = opxl.load_workbook('/Users/garethheyes/Downloads/ExcelImport2.xlsx')
    return wb


# Get the table of contents page so we know what receiving, print the information, assuming correct
def getprinttoc(ws):
    print("_____________TOC Details____________________")
    for row in ws.iter_rows(min_row=3, max_col=1, max_row=13):
        for cell in row:
            cv = cell.value
            print(cv)
            if "Declare-DataSheet" in cv or "Declare-header" in cv or "Declare-Data" in cv:
                # note I split the item before loading
                toclst.append(cell.value.split('>', 1)[-1])
    print("_________________________________________")
    return toclst


def returnheader(ws, headerrange):
    for item in headerrange:
        item = item.split('.', 1)[-1][:-1]

        if item.endswith('>'):
            item = item[:-1]
        else:
            item = item
        if len(item) > 1:
            colhead.append(ws[item].value)

    return colhead


# merge with print header when robust and works. Might make sense to change the library used though.
# not very elegant
def returndata(ws, datarange):
    for item in datarange:
        item = item.split('.', 1)[-1][:-1]

        if item.endswith('>'):
            item = item[:-1]
        elif item.endswith(')'):
            item = item[:-2]
        else:
            item = item
        if len(item) > 1:
            datalst.append(ws[item].value)
    return datalst


def chunks(listy, numy):
    # For item i in a range that is a length of l,
    for i in range(0, len(listy), numy):
        # Create an index range for l of n items:
        yield listy[i:i + numy]


# Gets the excel file note later on this will allow input
wb = getfile()
print(wb)
# prints the TOC and returns the items required for the data sheet

wst = wb['toc']
valuelist = getprinttoc(wst)

# get the items from the list and populate the variable
# note it seems to read sheets in lower case
datasheet = str(itemgetter(0)(valuelist)).lower()
ws = wb[datasheet]



datarange = str(itemgetter(2)(valuelist)).strip()
headerrange = str(ws[str(itemgetter(1)(valuelist))]).split(",")
#print(headerrange)
datarange = str(ws[str(datarange)]).split(",")
#print(datarange)


# call the print items

colprint = returnheader(ws, headerrange)
print('The column headers are:')
print(colprint)

#This returns the items from the data range
datalist = returndata(ws, datarange)

#This chunks up the data so that it prints more elegantly
finalListy = list(chunks(datalist, len(colhead)))

i = 0
while i < len(finalListy):
    print(finalListy[i])
    i += 1
